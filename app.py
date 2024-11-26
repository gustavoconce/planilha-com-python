import openpyxl

#Criar planilha
bd = openpyxl.Workbook()

#Visualizando páginas
print(bd.sheetnames)

#Criando página
bd.create_sheet('Frutas')

#Selecionando pagina
frutas_page = bd['Frutas']

#Adicionando dados
frutas_page.append(['Fruta', 'Qtd', 'Valor'])
frutas_page.append(['Banana', '5', 'R$3.90'])
frutas_page.append(['Maça', '12', 'R$9.10'])
frutas_page.append(['Melancia', '2', 'R$11,80'])
frutas_page.append(['Goiaba', '20', 'R$13.90'])
frutas_page.append(['Melão', '7', 'R$6.54'])

#Salvar planilha
bd.save('PlanilhaCompras.xlsx')