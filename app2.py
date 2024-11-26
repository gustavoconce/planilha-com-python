import openpyxl

#Importando planilha existente
bd = openpyxl.load_workbook('PlanilhaCompras.xlsx')

#Acessar pagina
frutas_page = bd['Frutas']

#Imprimindo os dados
#for rows in frutas_page.iter_rows(min_row = 2, max_row = 5):
#    print(rows[0].value, rows[1].value, rows[2].value)


#Alterando dados
for rows in frutas_page.iter_rows(min_row = 2, max_row = 5):
   for cell in rows:
      if cell.value == 'Banana':
         cell.value == 'Mamão'


#Salvar altreração
bd.save('PlanilhaCompras.xlsx')